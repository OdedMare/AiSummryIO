#!/usr/bin/env python
"""Ask the summaries API the same questions about every area in a spreadsheet.

One sheet row holds one area, as WKT (``MULTIPOLYGON``/``POLYGON``) or as a
GeoJSON MultiPolygon. For each row the script runs the questions configured
below through ``POST /api/v1/summaries``, waits for the run to finish, and
writes the answer back into the row's own columns. The workbook is saved after
every row, so an interrupted batch keeps everything it already earned and a
re-run resumes where it stopped.

Runs are sequential and separated by a cooldown (2 minutes by default) because
each one fans out to FLAPI packages and to the model; the point is to be a
polite client of a service the UI is also using.

Configure it by editing the CONFIG block below — questions, columns, cooldown.
Every value there can also be overridden for one run from the command line;
``--help`` lists them.

    cd backend
    pip install openpyxl            # not a runtime dependency of the service
    python scripts/area_batch.py --workbook areas.xlsx --dry-run
    python scripts/area_batch.py --workbook areas.xlsx
"""

import argparse
import json
import pathlib
import sys
import time
from typing import Any, Dict, List

import httpx

# Run from anywhere: the script lives beside the package it imports from.
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

from app.common.geometry import wkt_to_multipolygon  # noqa: E402

# --------------------------------------------------------------------------
# CONFIG — edit this block. Nothing below it needs changing for ordinary use.
# --------------------------------------------------------------------------

#: Where the API lives. Behind the nginx proxy this is "https://host/api/v1".
BASE_URL = "http://localhost:8000/api/v1"

#: The spreadsheet to read areas from and write answers into.
WORKBOOK = "areas.xlsx"

#: Sheet name, or "" for the active sheet.
SHEET = ""

#: Row holding the column headers. Set to 0 for a sheet with no header row —
#: every column below is then a column letter ("B", "C") instead of a title.
HEADER_ROW = 1

#: Column holding the area: WKT or a GeoJSON MultiPolygon, one area per row.
GEOMETRY_COLUMN = "אזור"

#: Optional column with an identifier to scope the run alongside the area.
#: Leave "" to run on the area alone.
ROOT_ID_COLUMN = ""

#: The questions to ask about every area, in order.
#:
#: - ``question``    the text sent to the agent.
#: - ``skill_keys``  up to 3 published Skill keys, or [] for none.
#: - ``outputs``     column title (or letter) -> field of the answer.
#:
#: Available fields: ``headline``, ``summary``, ``coverage``, ``key_findings``,
#: ``risks``, ``missing_data``, ``suggested_questions``, ``skill_results``,
#: plus ``status``, ``error``, ``run_id`` and ``conversation_id`` about the run
#: itself. Dotted paths reach inside: ``skill_results.0.summary``. Lists are
#: written as bulleted lines and anything else as JSON.
#:
#: A column named here is created at the end of the sheet if it is missing.
QUESTIONS = [
    {
        "question": "מה מאפיין את האזור הזה?",
        "skill_keys": [],
        "outputs": {
            "תשובה": "headline",
            "סיכום": "summary",
            "ממצאים": "key_findings",
            "סטטוס": "status",
        },
    },
    {
        "question": "אילו סיכונים עולים מהנתונים באזור?",
        "skill_keys": [],
        "outputs": {
            "סיכונים": "risks",
            "נתונים חסרים": "missing_data",
        },
    },
]

#: Seconds to wait between API calls. The wait happens *before* each call
#: except the first, so a batch never ends on an idle sleep.
COOLDOWN_SECONDS = 120

#: Ask the follow-up questions inside the first question's conversation, so
#: they reuse its evidence instead of re-running every package. Set to False
#: to give each question its own fresh conversation.
REUSE_CONVERSATION = True

#: Give up on a single run after this long and record the timeout in the row.
RUN_TIMEOUT_SECONDS = 900

#: Leave rows alone when their answer columns are already filled, which is
#: what makes a re-run resume rather than repeat.
SKIP_ANSWERED = True

# --------------------------------------------------------------------------

#: Statuses the backend treats as final; anything else means keep polling.
_FINISHED = ("completed", "partial", "failed")
_POLL_SECONDS = 5.0
#: The server blocks for at most 300s per request (MAX_WAIT_SECONDS).
_MAX_SERVER_WAIT = 300
#: Excel refuses a longer cell value.
_CELL_LIMIT = 32767


class SummaryClient:
    """The three calls this script makes, over one cookie session.

    One ``httpx.Client`` for the whole batch on purpose: the API identifies a
    conversation by a signed session cookie, and a follow-up is refused for a
    conversation belonging to another session. Sharing the client keeps every
    row's conversation reachable.
    """

    def __init__(self, base_url: str, timeout: float):
        self._base_url = base_url.rstrip("/")
        self._timeout = timeout
        self._http = httpx.Client(timeout=_MAX_SERVER_WAIT + 30)

    def close(self) -> None:
        self._http.close()

    def create(
        self, question: str, boundaries: Dict[str, Any],
        root_id: str, skill_keys: List[str],
    ) -> Dict[str, Any]:
        """Open a conversation on this area and run the first question."""
        payload = {
            "question": question,
            "skill_keys": skill_keys,
            "boundaries": boundaries,
        }
        if root_id:
            payload["root_id"] = root_id
        body = self._post("/summaries", payload)
        return {
            "conversation_id": body["conversation"]["id"],
            "run": self._await_run(body["run"]),
        }

    def follow_up(
        self, conversation_id: str, question: str, skill_keys: List[str],
    ) -> Dict[str, Any]:
        """Ask inside an open conversation, reusing its saved evidence."""
        run = self._post(
            "/conversations/%s/messages" % conversation_id,
            {"question": question, "skill_keys": skill_keys},
        )
        return {
            "conversation_id": conversation_id,
            "run": self._await_run(run),
        }

    def _post(self, path: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        response = self._http.post(
            self._base_url + path,
            json=payload,
            params={"wait": min(self._timeout, _MAX_SERVER_WAIT)},
        )
        if response.status_code >= 400:
            raise RuntimeError(
                "%s %s: %s" % (path, response.status_code, _detail(response))
            )
        return response.json()

    def _await_run(self, run: Dict[str, Any]) -> Dict[str, Any]:
        """Poll until the run finishes, past the server's own bounded wait."""
        deadline = time.monotonic() + self._timeout
        while run["status"] not in _FINISHED:
            if time.monotonic() >= deadline:
                run = dict(run)
                run["status"] = "timeout"
                run["error"] = "הריצה לא הסתיימה בתוך %ds" % self._timeout
                return run
            time.sleep(_POLL_SECONDS)
            response = self._http.get(self._base_url + "/runs/" + run["id"])
            if response.status_code >= 400:
                raise RuntimeError(
                    "GET /runs/%s: %s" % (run["id"], _detail(response))
                )
            run = response.json()
        return run


def _detail(response) -> str:
    try:
        return str(response.json().get("detail", response.text))
    except ValueError:
        return response.text[:200]


class Sheet:
    """Column lookup by header title, or by letter when there is no header."""

    def __init__(self, worksheet, header_row: int):
        self._sheet = worksheet
        self._header_row = header_row
        self._columns = {}
        if header_row:
            for cell in worksheet[header_row]:
                value = cell.value
                title = "" if value is None else str(value).strip()
                if title and title not in self._columns:
                    self._columns[title] = cell.column

    @property
    def first_data_row(self) -> int:
        return self._header_row + 1

    @property
    def last_row(self) -> int:
        return self._sheet.max_row

    def column(self, spec: str, create: bool = False) -> int:
        """Resolve a configured column to its 1-based index."""
        from openpyxl.utils import column_index_from_string

        if not self._header_row:
            return column_index_from_string(spec.strip().upper())
        title = spec.strip()
        if title in self._columns:
            return self._columns[title]
        if not create:
            raise KeyError("אין עמודה בשם %r בגיליון" % title)
        index = self._sheet.max_column + 1
        self._sheet.cell(self._header_row, index).value = title
        self._columns[title] = index
        return index

    def read(self, row: int, column: int) -> str:
        value = self._sheet.cell(row, column).value
        return "" if value is None else str(value).strip()

    def write(self, row: int, column: int, value: str) -> None:
        from openpyxl.styles import Alignment

        cell = self._sheet.cell(row, column)
        cell.value = value
        # A summary is a paragraph; without wrapping the sheet is unreadable.
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def geometry_of(text: str) -> Dict[str, Any]:
    """Read one cell as a GeoJSON MultiPolygon, whichever way it was written.

    WKT is what a GIS export puts in a cell; GeoJSON is what the API takes and
    what someone pasting from the map panel would have. Both are accepted, and
    a bare Polygon becomes a one-polygon MultiPolygon.
    """
    if text.startswith("{"):
        parsed = json.loads(text)
        kind = parsed.get("type")
        if kind == "Polygon":
            return {
                "type": "MultiPolygon", "coordinates": [parsed["coordinates"]]
            }
        if kind == "Feature":
            return geometry_of(json.dumps(parsed.get("geometry") or {}))
        if kind != "MultiPolygon":
            raise ValueError("נתמך MultiPolygon בלבד")
        return parsed
    boundaries = wkt_to_multipolygon(text)
    if not boundaries:
        raise ValueError("אין גיאומטריה בתא")
    return boundaries


def answer_field(payload: Dict[str, Any], path: str) -> Any:
    """Resolve a dotted output path against a finished run's answer."""
    value = payload
    for part in path.split("."):
        if isinstance(value, dict):
            value = value.get(part)
        elif isinstance(value, list) and part.isdigit():
            index = int(part)
            value = value[index] if index < len(value) else None
        else:
            return None
        if value is None:
            return None
    return value


def as_cell(value: Any) -> str:
    """Render an answer field as something a spreadsheet cell can hold.

    A field that came back empty — no risks, no missing data — is written as
    "—" rather than left blank. Blank has to keep meaning "never asked", or a
    resumed batch would ask the same question forever.
    """
    if value is None:
        text = ""
    elif isinstance(value, str):
        text = value
    elif isinstance(value, bool):
        text = "כן" if value else "לא"
    elif isinstance(value, (int, float)):
        text = str(value)
    elif isinstance(value, list) and all(
        isinstance(item, str) for item in value
    ):
        text = "\n".join("• " + item for item in value)
    else:
        text = json.dumps(value, ensure_ascii=False, indent=2)
    if not text and value is not None:
        text = "—"
    if len(text) > _CELL_LIMIT:
        text = text[:_CELL_LIMIT - 20] + "\n… (נחתך)"
    return text


def run_row(
    client: SummaryClient, sheet: Sheet, row: int, config: Dict[str, Any],
    cooldown, log,
) -> int:
    """Ask every configured question about one row's area.

    Returns how many API calls it made.
    """
    boundaries = geometry_of(sheet.read(row, config["geometry_column"]))
    root_id = (
        sheet.read(row, config["root_id_column"])
        if config["root_id_column"] else ""
    )
    conversation_id = ""
    calls = 0
    for index, question in enumerate(config["questions"]):
        targets = question["columns"]
        if config["skip_answered"] and all(
            sheet.read(row, column) for column in targets
        ):
            log("  שורה %d · שאלה %d — כבר נענתה, מדלג" % (row, index + 1))
            continue
        cooldown()
        calls += 1
        log("  שורה %d · שאלה %d: %s" % (row, index + 1, question["question"]))
        started = time.monotonic()
        try:
            if conversation_id and config["reuse_conversation"]:
                result = client.follow_up(
                    conversation_id, question["question"],
                    question["skill_keys"],
                )
            else:
                result = client.create(
                    question["question"], boundaries, root_id,
                    question["skill_keys"],
                )
            conversation_id = result["conversation_id"]
            run = result["run"]
        except (RuntimeError, httpx.HTTPError) as error:
            # One row's failure is not the batch's: record it and move on.
            run = {"id": "", "status": "failed", "error": str(error)}
        payload = dict(run.get("result") or {})
        payload.update({
            "status": run.get("status", ""),
            "error": run.get("error") or "",
            "run_id": run.get("id", ""),
            "conversation_id": conversation_id,
            "question": question["question"],
        })
        for column, path in targets.items():
            sheet.write(row, column, as_cell(answer_field(payload, path)))
        log("    %s ב-%ds" % (
            payload["status"], int(time.monotonic() - started)
        ))
    return calls


def make_cooldown(seconds: float, log):
    """Space out API calls, without a pointless sleep before the first one."""
    state = {"pending": False}

    def wait() -> None:
        if state["pending"] and seconds > 0:
            log("  המתנה %ds לפני הקריאה הבאה" % int(seconds))
            time.sleep(seconds)
        state["pending"] = True

    return wait


def build_config(args) -> Dict[str, Any]:
    questions = []
    for question in QUESTIONS:
        text = question["question"].strip()
        if not text:
            raise ValueError("שאלה ריקה בהגדרות")
        if len(question.get("skill_keys") or []) > 3:
            raise ValueError("אפשר לבחור עד 3 Skills לשאלה")
        questions.append({
            "question": text,
            "skill_keys": list(question.get("skill_keys") or []),
            "outputs": dict(question.get("outputs") or {}),
        })
    if not questions:
        raise ValueError("לא הוגדרו שאלות")
    return {
        "workbook": args.workbook,
        "out": args.out or args.workbook,
        "sheet": args.sheet,
        "header_row": args.header_row,
        "questions": questions,
        "cooldown": args.cooldown,
        "reuse_conversation": not args.new_conversation and REUSE_CONVERSATION,
        "skip_answered": SKIP_ANSWERED and not args.overwrite,
        "timeout": args.timeout,
    }


def parse_args(argv) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__.splitlines()[0],
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--workbook", default=WORKBOOK, help="קובץ ה-Excel")
    parser.add_argument(
        "--out", default="",
        help="קובץ יעד; ברירת המחדל היא לכתוב לאותו קובץ",
    )
    parser.add_argument("--sheet", default=SHEET, help="שם הגיליון")
    parser.add_argument("--header-row", type=int, default=HEADER_ROW)
    parser.add_argument("--base-url", default=BASE_URL)
    parser.add_argument(
        "--cooldown", type=float, default=COOLDOWN_SECONDS,
        help="שניות המתנה בין קריאות (ברירת מחדל %d)" % COOLDOWN_SECONDS,
    )
    parser.add_argument("--timeout", type=float, default=RUN_TIMEOUT_SECONDS)
    parser.add_argument(
        "--start-row", type=int, default=0, help="שורת הנתונים הראשונה לעיבוד",
    )
    parser.add_argument(
        "--limit", type=int, default=0, help="מספר שורות מרבי בריצה זו",
    )
    parser.add_argument(
        "--overwrite", action="store_true",
        help="הרץ גם שורות שכבר נענו",
    )
    parser.add_argument(
        "--new-conversation", action="store_true",
        help="שיחה נפרדת לכל שאלה במקום שאלות המשך",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="הצג מה יירוץ בלי לקרוא ל-API ובלי לשמור",
    )
    return parser.parse_args(argv)


def main(argv=None) -> int:
    from openpyxl import load_workbook

    args = parse_args(argv)
    config = build_config(args)

    def log(message: str) -> None:
        print(message, flush=True)

    # A mistyped column or sheet name is a config error, not a crash: say
    # which one, and stop before any run is charged for.
    try:
        workbook = load_workbook(config["workbook"])
        worksheet = (
            workbook[config["sheet"]] if config["sheet"] else workbook.active
        )
        sheet = Sheet(worksheet, config["header_row"])
        config["geometry_column"] = sheet.column(GEOMETRY_COLUMN)
        config["root_id_column"] = (
            sheet.column(ROOT_ID_COLUMN) if ROOT_ID_COLUMN else 0
        )
        for question in config["questions"]:
            question["columns"] = {
                sheet.column(column, create=True): path
                for column, path in question["outputs"].items()
            }
    except (KeyError, ValueError, OSError) as error:
        log("הגדרה שגויה: %s" % error)
        return 2

    rows = [
        row for row in range(
            max(args.start_row or sheet.first_data_row, sheet.first_data_row),
            sheet.last_row + 1,
        )
        if sheet.read(row, config["geometry_column"])
    ]
    if args.limit:
        rows = rows[:args.limit]
    log("%d שורות עם אזור, %d שאלות לכל שורה, המתנה של %ds בין קריאות" % (
        len(rows), len(config["questions"]), int(config["cooldown"]),
    ))
    if args.dry_run:
        for row in rows:
            try:
                cell = sheet.read(row, config["geometry_column"])
                boundaries = geometry_of(cell)
            except ValueError as error:
                log("  שורה %d — אזור פסול: %s" % (row, error))
                continue
            log("  שורה %d — %d פוליגונים" % (
                row, len(boundaries["coordinates"])
            ))
        log("ריצה יבשה: לא בוצעה שום קריאה ולא נשמר דבר.")
        return 0

    client = SummaryClient(args.base_url, config["timeout"])
    cooldown = make_cooldown(config["cooldown"], log)
    processed = 0
    try:
        for row in rows:
            log("שורה %d מתוך %d" % (row, sheet.last_row))
            try:
                run_row(client, sheet, row, config, cooldown, log)
            except ValueError as error:
                # A malformed area is this row's problem alone.
                log("  דילוג על שורה %d: %s" % (row, error))
                continue
            # Saved per row so an interrupted batch keeps its answers.
            workbook.save(config["out"])
            processed += 1
    except KeyboardInterrupt:
        log("הופסק. שומר את מה שהושלם.")
        workbook.save(config["out"])
        return 130
    finally:
        client.close()
    log("הסתיים: %d שורות נכתבו אל %s" % (processed, config["out"]))
    return 0


if __name__ == "__main__":
    sys.exit(main())
