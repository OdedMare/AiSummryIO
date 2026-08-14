"""Excel import/export for the Evaluation HTTP boundary."""

import json
import re
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_MAX_FILE_BYTES = 20 * 1024 * 1024
_MAX_ROWS = 10000
_CELL_LIMIT = 32767


def read_root_ids(data: bytes) -> dict:
    if not data:
        raise ValueError("קובץ ה-Excel ריק")
    if len(data) > _MAX_FILE_BYTES:
        raise ValueError("קובץ ה-Excel גדול מ-20MB")
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as error:
        raise ValueError("לא ניתן לקרוא את קובץ ה-Excel: %s" % error)
    matches = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30)):
            for cell in row:
                if _header(cell.value) == "rootid":
                    matches.append((sheet, cell.row, cell.column, str(cell.value)))
    if not matches:
        raise ValueError("לא נמצאה עמודת root_id בקובץ")
    if len(matches) > 1:
        locations = ", ".join(
            "%s!%s" % (sheet.title, get_column_letter(column))
            for sheet, _row, column, _value in matches[:5]
        )
        raise ValueError("נמצאו כמה עמודות root_id: %s" % locations)
    sheet, header_row, column, title = matches[0]
    root_ids, numeric = [], 0
    for row in range(header_row + 1, sheet.max_row + 1):
        cell = sheet.cell(row, column)
        value = _cell_identifier(cell.value, cell.number_format)
        if not value:
            continue
        if isinstance(cell.value, (int, float)):
            numeric += 1
        root_ids.append(value)
        if len(root_ids) > _MAX_ROWS:
            raise ValueError("אפשר להעלות עד 10,000 מזהים")
    if not root_ids:
        raise ValueError("עמודת root_id אינה מכילה מזהים")
    warnings = []
    if numeric:
        warnings.append(
            "%d מזהים נקראו כמספרים; מומלץ לעצב מזהים כטקסט כדי לשמור אפסים מובילים"
            % numeric
        )
    return {
        "root_ids": root_ids,
        "sheet": sheet.title,
        "column": title,
        "warnings": warnings,
    }


def evaluation_workbook(batch: dict, rows: List[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evaluation"
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A2"
    headers = [
        "root_id", "status", "headline", "summary", "coverage",
        "key_findings", "risks", "missing_data", "error", "rating",
        "comment", "duration_seconds", "run_id", "batch_label", "question",
        "skills", "cooldown_seconds",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="B8563C")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        result = row.get("result") or {}
        sheet.append([
            _excel_text(row["root_id"]), _excel_text(row["status"]),
            _excel_text(result.get("headline", "")),
            _excel_text(result.get("summary", "")),
            _excel_text(result.get("coverage", "")),
            _excel_text(result.get("key_findings")),
            _excel_text(result.get("risks")),
            _excel_text(result.get("missing_data")),
            _excel_text(row.get("error") or ""), row.get("rating"),
            _excel_text(row.get("comment") or ""),
            row.get("duration_seconds"), _excel_text(row.get("run_id") or ""),
            _excel_text(batch["label"]), _excel_text(batch["question"]),
            _excel_text(", ".join(batch.get("skill_keys") or [])),
            batch["cooldown_seconds"],
        ])
    sheet.auto_filter.ref = sheet.dimensions
    widths = [28, 14, 34, 70, 38, 50, 45, 45, 45, 10, 45, 16, 38, 24, 48, 28, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    target = BytesIO()
    workbook.save(target)
    return target.getvalue()


def _header(value: Any) -> str:
    return re.sub(r"[\s_-]+", "", str(value or "").strip().lower())


def _cell_identifier(value: Any, number_format: str) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        zeros = number_format.strip()
        if zeros and set(zeros) == {"0"}:
            return str(value).zfill(len(zeros))
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _excel_text(value: Any) -> str:
    if value is None:
        text = ""
    elif isinstance(value, list) and all(isinstance(item, str) for item in value):
        text = "\n".join("• " + item for item in value)
    elif isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, indent=2)
    else:
        text = str(value)
    text = text if len(text) <= _CELL_LIMIT else text[:_CELL_LIMIT - 12] + "\n… (נחתך)"
    # Excel interprets user-controlled text that begins with one of these
    # characters as a formula. Prefixing an apostrophe keeps the displayed
    # value intact while forcing a literal cell in exported workbooks.
    return "'" + text if text[:1] in ("=", "+", "-", "@") else text
